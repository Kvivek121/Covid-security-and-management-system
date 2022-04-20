from gpiozero import MCP3008
from smbus2 import SMBus
from mlx90614 import MLX90614
import RPi.GPIO as GPIO
import os
import max30100
import time
import Adafruit_CharLCD as lcd
import openpyxl

#import oxymeter_console

GPIO.setmode(GPIO.BCM)

pot = MCP3008(0) #potentiometer at analog pin 0 of MCP3008

#ultrasonic sensor pins

TRIG = 23 #16-board
ECHO = 24 #18- board

#motor pin L293D

#i_motor=18      #12-board
Motor1_in = 16  # Input Pin
Motor2_in = 12   # Input Pin
Motor3_en = 20  # Enable Pin

#setting up the gpio pins for l293d

GPIO.setup(Motor1_in,GPIO.OUT)
GPIO.setup(Motor2_in,GPIO.OUT)
GPIO.setup(Motor3_en,GPIO.OUT)

# IR sensor pin

ir_sensor = 17 #11-BOARD

# pulse sensor pin -vcc

Pulse_sensor = 27  #13-BOARD

#interfacing 16x2 LCD display

lcd_rs        = 5  
lcd_en        = 6
lcd_d4        = 13
lcd_d5        = 19
lcd_d6        = 26
lcd_d7        = 21

#creating object for LCD diplay
lcd_columns = 16
lcd_rows    = 2

#l=lcd.Adafruit_CharLCD(l1,l2,l3,l4,l5,l6,0,16,2)
l = lcd.Adafruit_CharLCD(lcd_rs, lcd_en, lcd_d4, lcd_d5, lcd_d6, lcd_d7,
                           lcd_columns, lcd_rows)

#setting of the GPIO pins as input/output

#GPIO.setup(i_motor,GPIO.OUT)
GPIO.setup(TRIG,GPIO.OUT)
GPIO.setup(ECHO,GPIO.IN)
GPIO.setup(ir_sensor,GPIO.IN)
GPIO.setup(Pulse_sensor,GPIO.OUT)

GPIO.output(Pulse_sensor,False)

GPIO.output(TRIG, False)

print ("Calibrating.....")
time.sleep(0.5)

print ("Place the object......")
l.clear()
s_lcd=str("     WELCOME   ")
l.message(s_lcd)

try:
    while True:
       
       sensor_data_list=[]
       #initializing pulse oximeter sensor
       
       mx30 = max30100.MAX30100()
       mx30.enable_spo2()
       
       #ultrasonic sensor initializing 
       
       GPIO.output(TRIG, GPIO.HIGH)
       time.sleep(0.00001)
       GPIO.output(TRIG, GPIO.LOW)

       while GPIO.input(ECHO)==0:
            pulse_start = time.time()

       while GPIO.input(ECHO)==1:
            pulse_end = time.time()

       pulse_duration = pulse_end - pulse_start
       distance = pulse_duration * 17150
       distance = round(distance+1.15, 2)

       if distance<=20 and distance>=5:
          print ("distance:",distance,"cm")
          print("motor on")
          #GPIO.output(i_motor,GPIO.HIGH)
          GPIO.output(Motor1_in,GPIO.HIGH)
          GPIO.output(Motor2_in,GPIO.LOW)
          GPIO.output(Motor3_en,GPIO.HIGH)
          l.clear()
          s_lcd=str("    Sanitized")
          l.message(s_lcd)
          time.sleep(2)
          
          #analog value through potentiometer 
          
          if (pot.value < 0.002):
              #led.value = 0
              print("Analog value",pot.value)
          else:
              #led.value = pot.value
              print("Analog value",pot.value)
          time.sleep(5*(pot.value)+0.001)
          print ("motor off.....")
          #GPIO.output(i_motor,GPIO.LOW)
          GPIO.output(Motor1_in,GPIO.LOW)
          GPIO.output(Motor2_in,GPIO.LOW)
          GPIO.output(Motor3_en,GPIO.LOW)
          
          #temnperature sensor values cal.
          
          list_of_amb_temperature=[]
          list_of_obj_temperature=[]
          
          try:
              for i in range(10):
                  bus = SMBus(1)
                  sensor = MLX90614(bus, address=0x5A)
                  list_of_amb_temperature.append(sensor.get_amb_temp())
                  list_of_obj_temperature.append(sensor.get_obj_temp())
             
                  time.sleep(0.1)
                  bus.close()
                  
              sum_obj_temperature=0
              
              for i in list_of_obj_temperature:
                  sum_obj_temperature+=i
              
              
              sensor_data_list.append(9/5*(sum_obj_temperature/10)+32)
              
              print("object_temperature",9/5*(sum_obj_temperature/10)+32)
              l.clear()
              s_lcd="Temperature "+ str(9/5*(sum_obj_temperature/10)+32)+"F"
              l.message(s_lcd)
              time.sleep(3)
              
          except:
              print("error")
         # try:
          time.sleep(3)
          if(GPIO.input(ir_sensor)==True): #object is far away
              print("NOT DETECTED")
                
            
          if(GPIO.input(ir_sensor)==False): #object is near
              print("DETECTED")
                
                #pulse sensor 
                
                #start=time.time()
                
              def moving_average(numbers):
                        window_size = 9
                        i = 0
                        # moving_averages = []
                        while i < len(numbers) - window_size + 1:
                            this_window = numbers[i : i + window_size]
                            window_average = sum(this_window) / window_size
                            # moving_averages.append(window_average)
                            i += 1
                        try:
                            return int((window_average/100))
                        except:
                            pass

                    # If HeartRate is <10 function assumes Fingure Not present and will not show incorrect data
                    # Also If SpO2 readings goes beyond 100. It will be shown as 100.
              def display_filter(moving_average_bpm,moving_average_sp02):
                        try:
                            if(moving_average_bpm<10):
                                moving_average_bpm ='NA'
                                moving_average_sp02 = 'NA'
                            else:
                                if(moving_average_sp02>100):
                                    moving_average_sp02 = 100
                            return moving_average_bpm, moving_average_sp02
                        except:
                            return moving_average_bpm, moving_average_sp02

              #while 1:
              print("keep your finger on the sensorfor 30 seconds")
              l.clear()
              s_lcd=str("Keep Your Finger for 30 second")
              l.message(s_lcd)
              #time.sleep(1.5)
            
              
              bpm_list=[]
              spo2_list=[]
              sum1=0
              sum2=0
              t_end=time.time()+30
              while time.time()< t_end :
                            #print("hello",time.time())
                            mx30.read_sensor()
                            hb = int(mx30.ir / 100)
                            spo2 = int(mx30.red / 100)
                            if mx30.ir != mx30.buffer_ir :
                                moving_average_bpm = (moving_average(mx30.buffer_ir))
                                # print(" MAX30.ir " + str(mx30.ir)) 
                                # print(" mx30.buffer_ir " + str(mx30.buffer_ir)) 
                                # print("|||||| Avg Pulse :" + str(moving_average_bpm)) 
                                # print("|||||| Pulse     :" + str(hb));
                            if mx30.red != mx30.buffer_red:
                                moving_average_sp02 = (moving_average(mx30.buffer_red))
                                # print(" MAX30.red " + str(mx30.red)) 
                                # print(" mx30.buffer_red " + str(mx30.buffer_red)) 
                                # print("###### Avg SpO2  :" + str(moving_average_sp02)) 
                                # print(" ##### SPO2     :" + str(spo2));
                            bpm,spo2 = display_filter(moving_average_bpm,moving_average_sp02)
                            
                            if bpm!=None and spo2!=None and bpm!='NA' and spo2!='NA'and bpm>=45 and spo2>=60:
                                bpm_list.append(bpm)
                                spo2_list.append(spo2)
                                time.sleep(1)
                            #print("hello")
                            #print(bpm_list)
                            sum1=sum(bpm_list)
                            sum2=sum(spo2_list)
              try:
                  avg_bpm=sum1/len(bpm_list)
                  avg_Spo2=sum2/len(bpm_list)
                  print(" *******************")
                  print(" bpm : "+ str(sum1/len(bpm_list)))
                  print(" SpO2: "+ str(sum2/len(bpm_list)))
                  print(" -------------------")
                  
                  sensor_data_list.append(avg_bpm)
                  sensor_data_list.append(avg_Spo2)
                  
                  l.clear()
                  s_lcd="Heart Rate "+str(avg_bpm)
                  l.message(s_lcd)
                  time.sleep(3)
                  
                  l.clear()
                  s_lcd="Spo2 "+ str(avg_Spo2)
                  l.message(s_lcd)
                  time.sleep(3)
                  l.clear()
                  
              except:
                  l.clear()
                  print("you didn't place the finger")
                  s_lcd=str("ops! Finger Not Placed")
                  l.message(s_lcd)
                  time.sleep(3)
                  l.clear()
                                
                        
                       
                    #oxymeter_console.display_filter()
                    
    #         except:
    #             print("pulse_error")
          if len(sensor_data_list)!=3:
              sensor_data_list.append(0.0)
              sensor_data_list.append(0.0)
          wb_obj = openpyxl.load_workbook("/home/pi/Workshop/FINAL/appending.xlsx")
          sheet_obj = wb_obj.active

          max_col = sheet_obj.max_column
          max_row = sheet_obj.max_row

          file=[]
          row_list=[]

        # Will all row value
          for j in range(1,max_row+2):

            file.append(row_list)
            row_list=[]

            for i in range(1, max_col + 1):
                cell_obj = sheet_obj.cell(row=j, column = i)
                row_list.append(cell_obj.value)
                #print(cell_obj.value, end = " ")

          #print(file)
          wb = openpyxl.Workbook()

          sheet = wb.active  
          #sheet['A1'] = 'Temperature' 
          #sheet['B1'] = 'Heart Rate' 
          #sheet['C1'] ='Spo2' 
          
          file.append(sensor_data_list)

          print(file)

          for i in file:
            if i!=[]:
                sheet.append(i)  
            
          wb.save('/home/pi/Workshop/FINAL/appending.xlsx')
          wb.close()
          l.clear() 
          s_lcd=str("     WELCOME   ")
          l.message(s_lcd)
                  
          if distance>20 and i==1:
              print ("place the object....")
        
          time.sleep(2)
          
          
          
except KeyboardInterrupt:
     GPIO.cleanup()
